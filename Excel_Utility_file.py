import openpyxl
from datetime import datetime

class ExcelUtil:

    def __init__(self, excel_path,sheet="test_data"):
        # Store Excel file path
        self.excel_path = excel_path
        # Open workbook (load Excel file into memory)
        self.workbook = openpyxl.load_workbook(self.excel_path)
        # gets sheet that has data
        self.sheet= self.workbook[sheet]

     # gets data for Pytest parametrize
    def get_param_data(self):
         # Reads all rows from Excel (skipping header row) & Returns list of tuples: [(row,username, password), ...]

        data_list = []  # []
         # Start from row 2 (skip header row)
        for row in range(2, self.sheet.max_row + 1):  # Start from row 2, skip headers
            username = self.sheet.cell(row=row, column=2).value # Column 2 = Username
            password = self.sheet.cell(row=row, column=3).value # Column 3 = Password
            data_list.append((row, username, password)) # Append tuple (row_no, username, password)

         # Close workbook after reading
        self.workbook.close()
        return data_list

    # Write test result
    def write_test_result(self,row, result, tester="Poornima"):
        """
        Writes test result back to Excel:
        - Date
        - Time
        - Tester name
        - Test status (Pass/Fail)
        """
        now = datetime.now()
        self.sheet.cell(row=row, column=4).value = now.strftime("%Y-%m-%d")   # Date
        self.sheet.cell(row=row, column=5).value = now.strftime("%H:%M:%S")   # Time
        self.sheet.cell(row=row, column=6).value = tester                     # Tester
        self.sheet.cell(row=row, column=7).value = result                     # Result

        # Save changes back to Excel
        self.workbook.save(self.excel_path)

        # Close workbook after writing
        self.workbook.close()


